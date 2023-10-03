#!/usr/bin/python
"""
    AWRp.py - Oracle AWR Parser 
    Copyright 2019 Alexandre Marti. All rights reserved. To more info send mail to alexandremarti@gmail.com.
    Licensed under the Apache License, Version 2.0. See LICENSE.txt for terms & conditions.

    Purpose: Print AWR HTML file content as a table format.
    Author: Alexandre Marti
                  
    Usage:       
        python AWRp.py -files <html file name list or mask>         

    Optional Parameters
        -fmt: table type output format. the default format is psql 
              csv,plain,simple,grid,fancy_grid,pipe,orgtbl,jira,presto,psql,rst,mediawiki,moinmoin,youtrack,html,latex,latex_raw,latex_booktabs,textile,excel
              more details on https://pypi.org/project/tabulate/
              
        -listsections : print a list of all sections parsed by AWRp.py. It´s useful to find section number for parameter -sections
        -sections: used to restrict which awr sections the AWRp.py will return. use -listsections to get the list of sections available.        

    Example1:
        python AWRp.py -files /files/awrfile.html
    
    Example2:    
        python AWRp.py -files /files/*.html
        
    Example3:    
        python AWRp.py -files '/files/awrfile01.html','/files/awrfile02.html'

    Example4:        
        python AWRp.py -listsections
        
    Example5:        
        python AWRp.py -files /files/*.html -sections 1,3,4               

    Help:
        python AWRp.py -h

"""

import pandas as pd
from bs4 import BeautifulSoup
from tabulate import tabulate
import glob   
import argparse
import os
from io import StringIO 
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Global variables        
TableSize = 10
NoneType = type(None)

def getData(content, filter):    
    '''
        Find AWR Table Section using the filter
    '''
    table = content.find('table', {'summary': filter})       
    # Testing if it didn´t find awr section
    if not isinstance(table, NoneType):    
        ret = pd.read_html(StringIO(str(table)), header=0)

        # 
        ret[0].rename(columns = {'Unnamed: 0': 'Info'}, inplace=True)
        return ret

def printTable(sectionIndex, sectionInfo, soup, fmt):  
    '''
        Print the pandas dataframe in the format defined
    '''
    data = getData(soup, sectionInfo)
    print("\nSECTION {}: {}\n".format(sectionIndex, sectionInfo))

    # for i in data:
    print('----')
    print(data)
    print('xxxx')

    return 

    df = data[0]

    if not isinstance(data, NoneType):
        match fmt:
            case 'csv':
                print(data[0].to_csv(index=False))
            case 'excel':
                if False:
                    outFileName = 'pandas_to_excel.xlsx'
                    if os.path.isfile(outFileName):
                        workbook = openpyxl.load_workbook(outFileName)
                    else:
                        workbook = openpyxl.Workbook()
                        workbook.save(outFileName)
                    worksheet = workbook.create_sheet('section' + str(sectionIndex))
                    worksheet.append([sectionInfo])
                    worksheet.append([])

                    for r in dataframe_to_rows(data[0], index=False, header=True):
                        worksheet.append(r)
                    workbook.save(outFileName)

                # update summary
                summaryDF = {}

                match sectionIndex:

                    case 1:
                        hostName = data[0].at[0, "Host Name"]                        
                        summaryDF.update({"hostName": hostName})
                        # print(summaryDF)

                        # print("Hostname: {}".format(hostName))
                        

                    case 2:                        
                        beginDateTime = data[0].at[0, "Snap Time"]
                        summaryDF.update({"beginDateTime": beginDateTime})

                        endDateTime = data[0].at[1, "Snap Time"]
                        summaryDF.update({"endDateTime": endDateTime})

                        beginSessions = data[0].at[0, "Sessions"]
                        summaryDF.update({"beginSessions": beginSessions})

                        endSessions = data[0].at[1, "Sessions"]
                        summaryDF.update({"endSessions": endSessions})

                        # print(summaryDF)
                        # print(beginDateTime, endDateTime, beginSessions, endSessions)

                    case 3:                        
                        logicalRead = data[0].at[4, "Per Second"]
                        summaryDF.update({"logicalRead": logicalRead})

                        physicalRead = data[0].at[6, "Per Second"]
                        summaryDF.update({"physicalRead": physicalRead})

                        physicalWrite = data[0].at[7, "Per Second"]
                        summaryDF.update({"physicalWrite": physicalWrite})

                        parsesSQL = data[0].at[15, "Per Second"]
                        summaryDF.update({"parsesSQL": parsesSQL})

                        executesSQL = data[0].at[20, "Per Second"]
                        summaryDF.update({"executesSQL": executesSQL})

                        print(summaryDF)
                        # print(logicalRead, physicalRead, physicalWrite, parsesSQL, executesSQL)

                    case 6:                        
                        beginLoadAverage = data[0].at[0, "Load Average Begin"]
                        endLoadAverage = data[0].at[0, "Load Average End"]
                        userCPU = data[0].at[0, "%User"]
                        print(beginLoadAverage, endLoadAverage, userCPU)

                    case 9: 
                        print(data[0])

                        topSQL = []
                        for i in range(TableSize):
                            topSQL.append(data[0].iloc[i])
                            # topSQL[i] = data[0].iloc[i + 1]
                        print(topSQL)


                # sales_sheet = workbook['section' + str(sectionIndex)]
                # with pd.ExcelWriter('pandas_to_excel.xlsx') as writer:                    
                    # writer.book = openpyxl.load_workbook('pandas_to_excel.xlsx')
                    # data[0].to_excel(writer, sheet_name='section' + str(sectionIndex))
                    # data[0].to_excel(writer, sheet_name='new_sheet2')

            case _:
                print(tabulate(data[0].iloc[0:TableSize], headers='keys', tablefmt=fmt, showindex=False))

                # Create an empty row with empty string values                
                #
                # empty_row = pd.DataFrame([empty_values], columns=df.columns)

    else:
        print('Data not found')

    ''' 
        if fmt == 'csv':            
            print(data[0].to_csv(index=False))
        else:
            print(tabulate(data[0], headers='keys', tablefmt=fmt, showindex=False) )  
    '''

def main():

    try:
        # CMD Parameters
        version = "1.0"
        version_description = "AWRp - Oracle AWR Parser {}".format(version)

        parser = argparse.ArgumentParser(description = version_description) 
        parser.add_argument("-files", help="Comma-delimited list of HTML AWR files", default="")
        parser.add_argument("-fmt", help="Table Format to display. The default is psql", 
                            default="psql", 
                            choices=['jira','csv','plain','simple','grid','fancy_grid','pipe','orgtbl','presto','psql','rst','mediawiki','moinmoin','youtrack','html','latex','latex_raw','latex_booktabs','textile','excel'])
        parser.add_argument("-listsections", help="List of all sections parsed by AWRp.py", action="store_true")        
        parser.add_argument("-sections", help="Comma-delimited numbers of awr sections to be returned by AWRp.py", default="")
               
        # If parse fail will show help
        args = parser.parse_args()
        
        sections = []
        if args.sections != "":
            sections = args.sections.split(",")

        # list of awr sections to parse
        infolist = ['This table displays database instance information', 
                    'This table displays host information',
                    'This table displays snapshot information',                     
                    'This table displays load profile', 
                    'This table displays top 10 wait events by total wait time', 
                    'This table displays wait class statistics ordered by total wait time',
                    'This table displays system load statistics',
                    'This table displays CPU usage and wait statistics',
                    'This table displays IO profile',
                    'This table displays top SQL by elapsed time', 
                    'This table displays top SQL by CPU time', 
                    'This table displays top SQL by user I/O time', 
                    'This table displays top SQL by buffer gets', 
                    'This table displays top SQL by physical reads', 
                    'This table displays top SQL by unoptimized read requests', 
                    'This table displays top SQL by number of executions', 
                    'This table displays top SQL by number of parse calls', 
                    'This table displays top SQL by amount of shared memory used', 
                    'This table displays top SQL by version counts', 
                    'This table displays top SQL by cluster wait time', 
                    'This table displays Foreground Wait Events and their wait statistics',
                    'This table displays top 10 wait events by total wait time'
                    ]
        
        if args.listsections:
            print("Printing the List of Parseable AWR Sections")
            print("===========================================\n")                
            for i, info in enumerate(infolist):                                
                print("    Section {}: {}".format(i, info)) 
                
        elif args.files == "":
            print("Ops! You need to pass the HTML file name.")
            print("Example:")            
            print("    AWRp.py -file /path/awrfile.html or AWRp.py -file /path/*.html")            
        else:                
            filelist = args.files.split(",")
            
            for fn in filelist:
                files=glob.glob(fn)           
                for file in files:
                    try:
                        print("Data from File: {}\n".format(file))
                        f=open(file)
                        
                        soup = BeautifulSoup(f,'lxml')     
                        for i, info in enumerate(infolist):                        
                            if len(sections) == 0 or str(i) in sections:
                                printTable(i, info, soup, args.fmt)

                    finally:
                        f.close()
    
    finally:    
        print("\nAWRp.py Finished\n")

if __name__ == '__main__':
    main()
    