from datetime import datetime

# import os
import pandas
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.styles import numbers
from openpyxl.formatting.rule import FormulaRule

import modules.constants as constants
import modules.configuration as configuration
import modules.arguments as arguments
import modules.common as common
import modules.aggregations as aggregations

from pathlib import Path

titles_row = 6

ns_yyyy_mm_dd_h_mm_ss = NamedStyle(name="datetime", number_format="yyyy-mm-dd h:mm:ss")
ns_h_mm_ss = NamedStyle(name="time", number_format="h:mm:ss")

redFill = PatternFill(start_color='EE1111', end_color='EE1111', fill_type='solid')
amberFill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')
greenFill = PatternFill(start_color='11EE11', end_color='11EE11', fill_type='solid')


# def create_sheet_checks(work_book: Workbook, tabs: list):
def create_sheet_checks(work_book: Workbook, df_sheets: pd.DataFrame, reports: dict):
    """
    Create the checks tab
    """
    work_sheet = work_book.create_sheet("Checks", 0)

    titles = [
        "tab",
        "file",
        "beginDateTime",
        "library_cache_lock_event_waits",
        "cursor_mutex_x_event_waits",
        "cursor_mutex_s_event_waits",
        "total_cpu",
        "version_one",
        "version_all",
        "library_cache_mutex_x_event_waits",
        "cursor_pin_s_wait_on_x_event_waits",
        "execute_to_parse",
        "concurrency_db_time",
        "ratio_exceptions_events"
    ]

    row = 2  # Start at row 2
    column = 2
    column = print_titles(titles, work_sheet, row=row, column=column)

    # row = 7
    # for j in range(0, len(tabs)):

    row = 3  # Start at row 2
    # For each of the tabs in the report (for each of the AWR files
    for row_tuple in df_sheets.itertuples(index=False):
        key = row_tuple[0]
        sheet = row_tuple[1]

        # Print Tab name
        column = 2
        work_sheet.cell(row=row, column=column, value=sheet)

        # For each of the columns
        column = 3
        for i in range(0, len(titles)):
            column_letter = get_column_letter(column)
            match column_letter:
                case "C":  # file
                    parameter = "file"
                    index_name = "Parameter"
                    column_name = "Value"
                    number_format: str = numbers.BUILTIN_FORMATS[3]
                    set_cell(reports=reports, key=key, section=constants.summary_section_key, index_name=index_name,
                             column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                             column=column, number_format=number_format)
                case "D":  # file
                    dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
                    parameter = "beginDateTime"
                    value = common.at(parameter, "Value", dataframe)
                    work_sheet.cell(row=row, column=column, value=value)
                    work_sheet.cell(row=row, column=column).style = ns_yyyy_mm_dd_h_mm_ss
                case "E":  # library_cache_lock_event_waits
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)
                    # print(f"dataframe: \n{dataframe}")

                    parameter = "library_cache_lock_event_waits"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "F":  # cursor_mutex_x_event_waits
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "cursor_mutex_x_event_waits"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "G":  # cursor_mutex_s_event_waits
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "cursor_mutex_s_event_waits"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "H":  # total_cpu
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "total_cpu"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)

                case "I":  # version_one
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "version_one"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "J":  # version_all
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "version_all"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "K":  # library_cache_mutex_x_event_waits
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "library_cache_mutex_x_event_waits"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "L":  # cursor_pin_s_wait_on_x_event_waits
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "cursor_pin_s_wait_on_x_event_waits"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "M":  # execute_to_parse
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "execute_to_parse"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)
                case "N":  # concurrency_db_time
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "execute_to_parse"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)

                case "O":  # ratio_exceptions_events
                    index_name = "Check"
                    section = constants.checks_section_key
                    dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

                    parameter = "ratio_exceptions_events"
                    column_name = "Result"
                    value = bool(common.at(parameter, column_name, dataframe))
                    work_sheet.cell(row=row, column=column, value=value)

                case _:
                    pass

            column += 1

        # Conditional formatting
        work_sheet.conditional_formatting.add(f"E{row}:O{row}", FormulaRule(formula=[
            f"E{row}=False"], stopIfTrue=True, fill=greenFill))
        work_sheet.conditional_formatting.add(f"E{row}:I{row}", FormulaRule(formula=[
            f"E{row}=True"], stopIfTrue=True, fill=redFill))
        work_sheet.conditional_formatting.add(f"J{row}:O{row}", FormulaRule(formula=[
            f"J{row}=True"], stopIfTrue=True, fill=amberFill))
        row += 1


def create_sheet_summary(work_book: Workbook, df_sheets: pd.DataFrame, reports: dict):

    titles = [
        "tab",
        "file",
        "hostName",
        "beginDateTime",
        "beginTime",
        "endDateTime",
        "elapsedTime",
        "dbTime",
        "dbTime %",
        "userCPU",
        "endSessions",
        "library_cache_lock_event_waits",
        "cursor_mutex_x_event_waits",
        "cursor_mutex_s_event_waits",
        "version_one",
        "version_all",
        "library_cache_mutex_x_event_waits",
        "cursor_pin_s_wait_on_x_event_waits",
        "execute_to_parse",
        "concurrency_db_time"
    ]

    """
    Create the summary tab
    """
    work_sheet = work_book.create_sheet("Summary", 1)

    # Print titles
    row = 2  # Start at row 2
    column = 2
    column = print_titles(titles, work_sheet, row=row, column=column)

    # Print tracked sql ids
    column = print_tracked_sql_ids(configuration.track_sql_ids, work_sheet, row=row, column=column)

    # Print aggregations
    column = print_tracked_sql_ids(aggregations.accum_aggregations.keys(), work_sheet, row=row, column=column)

    row = 3  # Start at row 2
    # For each of the tabs in the report (for each of the AWR files
    for row_tuple in df_sheets.itertuples(index=False):
        key = row_tuple[0]
        sheet = row_tuple[1]

        # Print Tab name
        # column = 2
        # work_sheet.cell(row=row, column=column, value=sheet)

        # For each of the columns
        column = 2
        for i in range(0, len(titles)):
            print_summary_defined_columns(sheet, column, key, reports, row, work_sheet)
            column += 1

        # Print tracked SQL Ids
        section = constants.tracked_sql_ids_section_key
        index_name = "SQL Id"
        dataframe = get_dataframe(reports, key, section=section, index_name=index_name)

        for track_sql_id in configuration.track_sql_ids:
            value = common.at(track_sql_id, "Executions", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).number_format = numbers.BUILTIN_FORMATS[3]
            column += 1

        # Print aggregations
        # for aggregation_key in aggregations.accum_aggregations.keys():
        #     value = aggregations.accum_aggregations.get(aggregation_key)
        #     work_sheet.cell(row=row, column=column, value=value)
        #     work_sheet.cell(row=row, column=column).number_format = numbers.BUILTIN_FORMATS[3]
        #     column += 1
        row += 1


def print_summary_defined_columns(sheet, column, key, reports, row, work_sheet):
    column_letter = get_column_letter(column)
    match column_letter:
        case "B":  # tab
            work_sheet.cell(row=row, column=column, value=sheet)
        case "C":  # file
            parameter = "file"
            index_name = "Parameter"
            column_name = "Value"
            number_format: str = numbers.BUILTIN_FORMATS[3]
            set_cell(reports=reports, key=key, section=constants.summary_section_key, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "D":  # hostName
            parameter = "hostName"
            index_name = "Parameter"
            column_name = "Value"
            number_format = numbers.BUILTIN_FORMATS[3]
            set_cell(reports=reports, key=key, section=constants.summary_section_key, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "E":  # beginDateTime
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "beginDateTime"
            value = common.at(parameter, "Value", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).style = ns_yyyy_mm_dd_h_mm_ss

        case "F":  # beginTime
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "beginTime"
            value = common.at(parameter, "Value", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).style = ns_h_mm_ss

        case "G":  # endDateTime
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "endDateTime"
            value = common.at(parameter, "Value", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).style = ns_yyyy_mm_dd_h_mm_ss

        case "H":  # elapsedTime
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "elapsedTime"
            value = common.at(parameter, "Value", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).number_format = "0.00"

        case "I":  # dbTime
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "dbTime"
            value = common.at(parameter, "Value", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).number_format = "0.00"

        case "J":  # dbTime %
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "elapsedTime"
            elapsedTime: float = common.at(parameter, "Value", dataframe)

            parameter = "dbTime"
            dbTime: float = common.at(parameter, "Value", dataframe)

            value = round(dbTime / elapsedTime, 2)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).number_format = "0.00"

        case "K":  # userCPU
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "userCPU"
            value = common.at(parameter, "Value", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).number_format = "0.00"

        case "L":  # endSessions
            dataframe = get_dataframe(reports, key, constants.summary_section_key, "Parameter")
            parameter = "endSessions"
            value = common.at(parameter, "Value", dataframe)
            work_sheet.cell(row=row, column=column, value=value)
            work_sheet.cell(row=row, column=column).number_format = numbers.BUILTIN_FORMATS[3]

        case "M":  # library_cache_lock_event_waits
            parameter = "library_cache_lock_event_waits"
            index_name = "Check"
            column_name = "Evidence"
            number_format = numbers.BUILTIN_FORMATS[3]
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "N":  # cursor_mutex_x_event_waits
            parameter = "cursor_mutex_x_event_waits"
            index_name = "Check"
            column_name = "Evidence"
            number_format = numbers.BUILTIN_FORMATS[3]
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "O":  # cursor_mutex_s_event_waits
            parameter = "cursor_mutex_s_event_waits"
            index_name = "Check"
            column_name = "Evidence"
            number_format = numbers.BUILTIN_FORMATS[3]
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "P":  # version_one
            parameter = "version_one"
            index_name = "Check"
            column_name = "Evidence"
            number_format = numbers.BUILTIN_FORMATS[3]
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "Q":  # version_all
            parameter = "version_all"
            index_name = "Check"
            column_name = "Evidence"
            number_format = numbers.BUILTIN_FORMATS[3]
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "R":  # library_cache_mutex_x_event_waits
            parameter = "library_cache_mutex_x_event_waits"
            index_name = "Check"
            column_name = "Evidence"
            number_format = numbers.BUILTIN_FORMATS[3]
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "S":  # cursor_pin_s_wait_on_x_event_waits
            parameter = "cursor_pin_s_wait_on_x_event_waits"
            index_name = "Check"
            column_name = "Evidence"
            number_format = numbers.BUILTIN_FORMATS[3]
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "T":  # execute_to_parse
            parameter = "execute_to_parse"
            index_name = "Check"
            column_name = "Evidence"
            number_format = "0.00"
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)

        case "U":  # concurrency_db_time
            parameter = "concurrency_db_time"
            index_name = "Check"
            column_name = "Evidence"
            number_format = "0.00"
            section = constants.checks_section_key

            set_cell(reports=reports, key=key, section=section, index_name=index_name,
                     column_name=column_name, parameter=parameter, work_sheet=work_sheet, row=row,
                     column=column, number_format=number_format)
        case _:
            pass


def set_cell(reports, key, section, index_name, column, column_name, number_format, parameter, row, work_sheet):
    dataframe = get_dataframe(reports, key, section, index_name)
    value = common.at(parameter, column_name, dataframe)
    work_sheet.cell(row=row, column=column, value=value)
    work_sheet.cell(row=row, column=column).number_format = number_format


def get_dataframe(reports, key, section, index_name):
    report = reports[key]
    data_frame = report[section]
    # print(f"summary_df: \n{summary_df} - \nindex: {summary_df.index.name}")
    if index_name != data_frame.index.name:
        data_frame.set_index(index_name, inplace=True)
    return data_frame


def print_titles(titles, work_sheet, row, column):
    for i in range(0, len(titles)):
        work_sheet.cell(row=row, column=column, value=titles[i])
        column += 1
    return column


def print_tracked_sql_ids(list, work_sheet, row, column):
    # Print the titles in the list
    for item in list:
        work_sheet.cell(row=row, column=column, value=item)
        column += 1
    return column


def print_reports(reports: dict):
    # print (reports) sorted by date
    list_keys = sorted(reports.keys())

    wb = Workbook()
    sheet_counter = 0
    sheets = []
    # write one sheet per AWR
    for key in list_keys:
        # Create a tab per AWR file
        sheet_name = f"AWR2Excel_{sheet_counter}"
        sheets.append(sheet_name)
        # sheets.append({"key": key, "sheet_name": sheet_name})
        # sheets.append({"key": key, "sheet_name": sheet_name})
        # print(f"sheets: {sheets}")

        sheet_counter += 1
        ws = wb.create_sheet(sheet_name)
        report = reports[key]

        # Write the sections of an AWR
        for section_key in report:
            section_df = report[section_key]
            ws.append([])

            ws.append([section_key])

            for row in dataframe_to_rows(section_df, index=False, header=True):
                ws.append(row)

    # name = sheets[0]["sheet_name"]
    # print(f"sheets: {name}")
    # d = {'keys': list_keys, 'sheets': sheets}
    df_sheets = pandas.DataFrame(data={'keys': list_keys, 'sheets': sheets})
    # print(df)

    # create checks sheet
    create_sheet_checks(wb, df_sheets, reports)

    # create summary sheet
    create_sheet_summary(wb, df_sheets, reports)

    # convert datetime obj to string
    # write a file object along with .xlsx extension
    str_current_datetime = str(datetime.now().strftime("%Y-%m-%d %H-%M-%S"))
    out_filename = str_current_datetime + "_AWR2Excel" + ".xlsx"

    args = arguments.get_args()
    output_path = Path(args.output)
    Path(output_path).mkdir(parents=True, exist_ok=True)
    wb.save(output_path / out_filename)
